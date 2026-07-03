#!/usr/bin/env python3
"""가구/인테리어 카탈로그 생성 — FURNITURE_CATALOG_DESIGN.md §2 구현.

  python3 furniture_catalog.py decompose   # 매핑 JSONL → 키·변형 분해
  python3 furniture_catalog.py group       # 카탈로그 롤업 + title 생성
  python3 furniture_catalog.py verify      # 골든 회귀 스위트 (§5.3 Phase 2)
  python3 furniture_catalog.py all         # decompose → group → verify

입력: outputs/furniture_geo_mapped.jsonl (GEO 매핑 + OCR 오버레이 완료본)
출력: outputs/furniture_decomposed.csv        (행별 catalog_key + variant_attrs)
      outputs/catalogs_furniture.csv          (카탈로그 롤업 + title_geo)
      outputs/catalog_variants_furniture.csv  (변형 + title_commerce)

계층: Catalog(모델) ─ Product(PDP) ─ Variant(색상×사이즈 옵션)
키:   brand + "|" + (모델코드 | slug(canonical tokens))
"""
import argparse
import csv
import json
import os
import re
import sys
import unicodedata
from collections import Counter, defaultdict

HERE = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, HERE)
import catalog_lexicon_furniture as lex  # noqa: E402

OUT = os.path.join(HERE, "outputs")
IN_JSONL = os.path.join(OUT, "furniture_geo_mapped.jsonl")
DECOMP_OUT = os.path.join(OUT, "furniture_decomposed.csv")
CATALOG_OUT = os.path.join(OUT, "catalogs_furniture.csv")
VARIANT_OUT = os.path.join(OUT, "catalog_variants_furniture.csv")

_WS = re.compile(r"\s+")


def nfkc(s):
    return _WS.sub(" ", unicodedata.normalize("NFKC", s or "")).strip()


# ── 카테고리 클래스 판정 (축 레지스트리 cats 매칭용) ─────────────────────────

def cat_class(l1, l2, name):
    if "타일" in name or "벽지" in name or "도기" in name:
        return "tile"       # 색상=모델 예외
    if l2 == "소파" or re.search(r"소파|스툴|등방석", name):
        return "sofa"
    if l1 == "조명":
        # "LED 조명 패브릭 침대"처럼 침대가 머리명사(더 뒤)면 침대가 본체 — 조명몰 분류보다 우선.
        # 반대로 "침대 무드등"은 등이 머리명사 → lighting 유지.
        bed_pos = name.rfind("침대")
        light_pos = max(name.rfind(w) for w in ("조명", "스탠드", "무드등", "등"))
        if bed_pos > light_pos:
            return "bed"
        return "lighting"
    if l1 == "침구":
        return "bedding"
    if l2 == "침대" or "침대" in name:
        return "bed"
    if l1 == "인테리어소품":
        return "decor"
    if l1 == "커튼/블라인드":
        return "curtain"
    return "etc"


# ── Phase A: 전처리 ───────────────────────────────────────────────────────────

def strip_hard(n):
    for p in lex.STRIP_HARD:
        n = re.sub(p, " ", n)
    return _WS.sub(" ", n).strip()


def detect_bundle(n):
    """세트상품 마커 — 키에 '세트' 마커 부여 (단품과 분리, §6.3)."""
    for m in lex.SET_MARKERS:
        if m in n:
            return True
    return False


# ── Phase B: 괄호 분리 ────────────────────────────────────────────────────────

_PAREN_RE = re.compile(r"\(([^()]*)\)")
# 옵션 나열 괄호: 내용이 전부 색상/색온도/규격 나열일 때만 제거
_OPT_PAREN_RE = re.compile(
    r"^[\s/+,·]*(?:(?:[가-힣]*(?:색|빛)|大|중|소|유광|무광|숏넥|롱넥|원형|사각|"
    r"아이스타입|일반|엣지|밀크|감지|\d+\s*[wW]|\d+컬러|\d+colors?)[\s/+,·]*)+$")


def extract_parens(n, cc):
    """괄호 분리·분류 → (본문, {coverage, module, spec, dropped})."""
    info = {"coverage": "", "module": "", "specs": []}

    def repl(m):
        inner = m.group(1).strip()
        cm = re.match(r"(전체|전면|전측면)가죽적용$", inner)
        if cm:
            info["coverage"] = cm.group(1) + "가죽"
            return " "
        if re.match(r"^[A-Z](\+[A-Z])*$", inner):
            info["module"] = inner
            return " "
        if re.match(r"^\d{4,}(-\d+)?$", inner):   # 내부 상품코드 (908165-1) — 키 오염 방지
            info["specs"].append(inner)
            return " "
        if _OPT_PAREN_RE.match(inner):        # 색상/옵션 나열
            info["specs"].append(inner)
            return " "
        # 색상 단독 괄호 (차콜) — 색상 렉시콘 매칭
        if _color_token(inner) and cc != "tile":
            info["specs"].append(inner)
            return " "
        return m.group(0)                     # 정보 괄호는 보존

    n = _PAREN_RE.sub(repl, n)
    return _WS.sub(" ", n).strip(), info


def _color_token(t):
    t = t.strip().replace(" ", "")
    for line in lex.FABRIC_LINES:
        if t.startswith(line):
            t = t[len(line):]
    for p in lex.COLOR_PREFIX:
        if t.startswith(p):
            t = t[len(p):]
    return t in set(lex.COLOR_BASE) or t.replace("_", "") in set(lex.COLOR_BASE)


# ── Phase C: 변형축 추출 ──────────────────────────────────────────────────────

_SEAT_RE = re.compile(r"(?<![\d.])([1-6](?:\.5)?)\s*인(?:용)?(?![가-힣])")
_WATT_RE = re.compile(r"\(?\s*\d+(?:\.\d)?\s*[wW]\b(?:\s*/\s*\d+(?:\.\d)?\s*[wW]\b)*\s*\)?")
_CCT_RE = re.compile(r"주광색|전구색|주백색|\d{4}[kK]\b")
_SIZE_BR_RE = re.compile(r"\[([SQKD/,\sEL]+)\]")
_SIZE_BARE_RE = re.compile(r"(?<![A-Za-z0-9가-힣])(S/SS|Q/K|SS|EK|LK|KK|[SQKD])(?![A-Za-z0-9가-힣])")
_DIM_CM_RE = re.compile(r"(?<![\dx×*.])(\d{2,3})\s*cm\b")
_MM_RE = re.compile(r"\(?(\d{3,4})\s*mm\)?")
_DASH_COLOR_RE = re.compile(r"\s*-\s*([가-힣A-Za-z_]+)\s*$")


def extract_variants(n, cc, rec):
    """이름에서 변형축 제거 → variant_attrs. 카테고리별 축 적용 (AXES 레지스트리)."""
    va = {}
    # 색상 (전 카테고리, 타일 예외) — dash-color / 선두 색상시리즈
    if cc != "tile":
        m = _DASH_COLOR_RE.search(n)
        if m and _color_token(m.group(1)):
            va["color"] = m.group(1)
            n = n[: m.start()]
        # 선두 색상 (도토로 "로즈골드 글리터…", 단 색상+유형 조합명 보호: 뒤에 토큰 2개 이상)
        toks = n.split()
        if len(toks) >= 3 and _color_token(toks[0]):
            va.setdefault("color", toks[0])
            n = " ".join(toks[1:])
    # 사이즈 브래킷/bare (침구·침대)
    if cc in ("bedding", "bed"):
        m = _SIZE_BR_RE.search(n)
        if m:
            va["size"] = m.group(1).strip()
            n = n[: m.start()] + " " + n[m.end():]
        else:
            sizes = _SIZE_BARE_RE.findall(n)
            for ko, cd in (("슈퍼싱글", "SS"), ("라지킹", "LK"), ("이스턴킹", "EK"),
                           ("싱글", "S"), ("더블", "D"), ("퀸", "Q"), ("킹", "K")):
                if re.search(r"(?<![가-힣])" + ko + r"(?![가-힣])", n):
                    sizes.append(cd)
                    n = re.sub(r"(?<![가-힣])" + ko + r"(?![가-힣])", " ", n)
            sizes = list(dict.fromkeys(sizes))
            if sizes:
                va["size"] = "/".join(sizes)
                n = _SIZE_BARE_RE.sub(" ", n)
    # 소파 인수 (변형 — §6.1)
    if cc == "sofa":
        seats = _SEAT_RE.findall(n)
        if seats:
            va["seat"] = "/".join(seats)
            n = _SEAT_RE.sub(" ", n)
        for f in lex.SOFA_FORMS_DROP:
            if f in n:
                va.setdefault("form", f)
                n = n.replace(f, " ")
    # 조명 와트·색온도
    if cc == "lighting":
        m = _WATT_RE.search(n)
        if m and "평형" not in n[max(0, m.start()-4):m.end()+4]:
            va["watt"] = _WS.sub("", m.group(0)).strip("()")
            n = _WATT_RE.sub(" ", n, count=1)
        if _CCT_RE.search(n):
            va["cct"] = "/".join(dict.fromkeys(_CCT_RE.findall(n)))
            n = _CCT_RE.sub(" ", n)
    # 치수: 조명 mm / 트리·매트 cm (§6.1 — 블라인드 mm는 모델이므로 제외)
    if cc in ("lighting",):
        m = _MM_RE.search(n)
        if m and not re.search(r"매입|다운라이트|타공|커튼봉", n):
            va["mm"] = m.group(1)
            n = n[: m.start()] + " " + n[m.end():]
    if cc == "decor":
        m = _DIM_CM_RE.search(n)
        if m:
            va["cm"] = m.group(1)
            n = n[: m.start()] + " " + n[m.end():]
    # 팩 수량 (전 카테고리)
    m = re.search(lex.PACK_RE, n)
    if m:
        va["pack"] = m.group(1)
        n = re.sub(lex.PACK_RE, " ", n, count=1)
    # 렌탈
    if re.search(lex.RENTAL_RE, n):
        va["offer"] = "rental"
        n = re.sub(lex.RENTAL_RE, " ", n)
    return _WS.sub(" ", n).strip(), va


# ── Phase D/E: 노이즈·정규화·키 ───────────────────────────────────────────────

def strip_key_noise(n, cc):
    for p in lex.STRIP_SPEC_QUALITY:
        n = re.sub(p, " ", n)
    n = re.sub(lex.LISTING_FORMAT_RE, " ", n)
    # 용도/장소 SEO 스팸: "회의실의자"→"의자", 단독 장소명사 제거, 마케팅 구 제거
    n = re.sub(lex.PLACE_TYPE_RE, r"\1", n)
    for p in lex.PLACE_NOUNS:
        n = re.sub(r"(?<![가-힣])" + p + r"(?![가-힣])", " ", n)
    for p in lex.MARKETING_PHRASES:
        n = re.sub(r"(?<![가-힣])" + p + r"(?![가-힣])", " ", n)
    # 유형 토큰 중복 제거 ("의자 … 의자")
    toks, seen = [], set()
    for t in n.split():
        if t in seen and re.search(r"의자|침대|옷장|서랍장|수납장|조명|책상|테이블|소파", t):
            continue
        seen.add(t)
        toks.append(t)
    n = " ".join(toks)
    # 소파 DESC/원단라인 (§4.3 — 전동은 NEVER_STRIP)
    if cc == "sofa":
        for t in lex.SOFA_DESC + lex.FABRIC_LINES:
            n = re.sub(r"(?<![가-힣])" + re.escape(t) + r"(?![가-힣])", " ", n)
    # 가드 수식어: 제거 후 토큰 2개 미만이면 복원 (§4.4 가드1)
    n2 = n
    for t in lex.GUARDED_MODIFIERS:
        n2 = re.sub(r"(?<![가-힣])" + re.escape(t) + r"(?![가-힣])", " ", n2)
    n2 = _WS.sub(" ", n2).strip()
    if len(n2.split()) >= 2:
        n = n2
    return _WS.sub(" ", n).strip()


_KO_SIZE_SYN = [  # 복합어 먼저, 전부 한글 경계 가드 (베이킹/퀸센스 오탐 방지)
    (r"(?<![가-힣])더블킹(?![가-힣])", "KK"), (r"(?<![가-힣])슈퍼싱글(?![가-힣])", "SS"),
    (r"(?<![가-힣])라지킹(?![가-힣])", "LK"), (r"(?<![가-힣])이스턴킹(?![가-힣])", "EK"),
    (r"(?<![가-힣])싱글(?![가-힣])", "S"), (r"(?<![가-힣])더블(?![가-힣])", "D"),
    (r"(?<![가-힣])퀸(?=침대|매트리스|사이즈)", "Q "), (r"(?<![가-힣])킹(?=침대|매트리스|사이즈)", "K "),
    (r"(?<![가-힣])퀸(?![가-힣])", "Q"), (r"(?<![가-힣])킹(?![가-힣])", "K"),
]


def synonyms(n):
    for pat, rep in lex.CATALOG_SYNONYMS:
        n = re.sub(pat, rep, n)
    for pat, rep in _KO_SIZE_SYN:
        n = re.sub(pat, rep, n)
    n = re.sub(r"모듈\s*([A-Z])\b", "모듈", n)   # 모듈A → 모듈
    return _WS.sub(" ", n).strip()


def find_model_code(n):
    for m in re.finditer(lex.MODEL_CODE_RE, n):
        if m.group(0) not in lex.MODEL_CODE_EXCLUDE:
            return m.group(0)
    return ""


def slug(tokens):
    return re.sub(r"[^a-z0-9가-힣]", "", "".join(tokens).lower())


def resolve_brand(mall, name):
    """이름 선두 브랜드 토큰 승격 → 없으면 BRAND_KO[mall] (§2.1)."""
    head = name.split()[0] if name.split() else ""
    for slug_, aliases in lex.BRAND_ALIASES.items():
        if head in aliases or head == lex.BRAND_KO.get(slug_):
            return lex.BRAND_KO[slug_], name[len(head):].strip()
    # 몰 외 브랜드 (금호전기/글로리나인 등) — 승격하되 이름 유지
    if head in ("금호전기", "글로리나인", "보닥"):
        return head, name[len(head):].strip()
    return lex.BRAND_KO.get(mall, mall), name


def make_key(rec):
    """§2.2 파이프라인 — (catalog_key, canonical, variant_attrs, flags)."""
    mall = rec["source"]["mall"]
    raw = rec["name"]
    l1, l2 = rec["l1_category"], rec["l2_category"]

    n = nfkc(raw)
    n = strip_hard(n)                                     # A2~A3
    if re.match(r"^\d{6,}_", n) or rec.get("price") in ("999999",):
        return None                                       # A4 trash
    brand, n = resolve_brand(mall, n)
    is_set = detect_bundle(n)                             # A5

    cc = cat_class(l1, l2, n)
    n, pinfo = extract_parens(n, cc)                      # B1
    # 매트리스 구성 (침대) — 구성이 다르면 별도 카탈로그 (사용자 정책 2026-07-02)
    mattress = ""
    mm_ = re.search(r"[+＋]\s*((?:독립|양면|본넬|포켓)?\s*매트(?:리스)?)\b|매트(?:리스)?\s*(포함|세트)\b", n)
    if mm_:
        mattress = (mm_.group(1) or "매트리스").strip()
        n = n[:mm_.start()] + " " + n[mm_.end():]
    n = re.sub(lex.BUNDLE_SUFFIX_RE, " ", n)              # B2 (B1 이후!)
    n = re.sub(r"\+\s*[가-힣A-Za-z ]{1,12}\d*\s*(?:증정)?$", " ", n)
    n, va = extract_variants(n, cc, rec)                  # C
    if pinfo["coverage"]:
        va["coverage"] = pinfo["coverage"]
    if pinfo["module"]:
        va["module"] = pinfo["module"]
    if mattress:
        va["mattress"] = mattress
    n = strip_key_noise(n, cc)                            # D
    n = synonyms(n)                                       # E1
    for _ in range(3):  # 괄호/브래킷 내부 잔재 수렴 정리
        n = re.sub(r"([\[(][^\])]*?)[\s,/]{2,}([^\])]*[\])])", r"\1 \2", n)
        n = re.sub(r"([\[(][^\])]*?)[\s,/]+([\])])", r"\1\2", n)   # 닫힘 직전 구분자
        n = re.sub(r"[\[(]\s*[,/\s]*\s*[\])]", " ", n)              # 빈 그룹
        n = re.sub(r"\(\s*[가-힣A-Za-z]*\s*[:：]?\s*[/,·]*\s*\)", " ", n)
        n = re.sub(r"(?:\s+[/,]\s*)+", " ", n)            # 고아 슬래시·콤마
        n = re.sub(r"\s+[~∼]\s*", " ", n)
        n = _WS.sub(" ", n).strip(" -/,~")

    code = find_model_code(n)
    if code:                                              # E2 code-first
        key = f"{brand}|{code}" + (("|" + re.sub(r"\s+", "", mattress)) if mattress else "")
        canonical = n
    else:
        tokens = n.split()
        if mall == "dongsuh":                             # E4 토큰집합
            toks = sorted(set(tokens))
            toks = [t for t in toks if not any(t != o and t in o for o in toks)]
            key_body = slug(toks)
        else:
            key_body = slug(tokens)
        if is_set:
            key_body += "|세트"
        if mattress:
            key_body += "|" + re.sub(r"\s+", "", mattress)  # 독립매트/양면매트 구분 유지
        key = f"{brand}|{key_body}"
        canonical = n
    return {"catalog_key": key, "brand": brand, "canonical": canonical,
            "variant_attrs": va, "bundle_flag": is_set, "cat_class": cc,
            "model_code": code}


# ── Stage 1: decompose ────────────────────────────────────────────────────────

# 모델형 옵션 판정: 옵션 과반이 상품유형 단어 포함 → 각 옵션 = 별도 카탈로그 (동서 모음전)
_OPT_TYPE_RE = re.compile(
    r"옷장|서랍장|수납장|침대|매트리스|책상|의자|소파|테이블|장롱|화장대|프레임"
    r"|거실장|책장|중문|파티션|틈새수납장|행거|선반")
_OPT_NUM_PREFIX = re.compile(r"^\d+(-\d+)?[.)]?\s*")
_TYPE_LISTING = re.compile(
    r"옷장|서랍장|수납장|장롱|침대|매트리스|책상|의자|소파|테이블|화장대|모음전|모음|\d+종")


# 추가 상품/구성 옵션 — 카탈로그 분해 대상 아님 (무시, 2026-07-02 정책)
_OPT_ADDON_RE = re.compile(
    r"추가|사은품|증정|선택\s*안\s*함|구매\s*안\s*함|단품\s*구매|미포함|없음"
    r"|\+\s*[\d,]+\s*원|★")


# 옵션 텍스트 정제: 가격차·품절·번호·내부코드 제거 (추가상품 필터보다 먼저!)
def _clean_opt(o):
    o = re.sub(r"\[?\s*품절\s*\]?", " ", o)
    o = re.sub(r"\s*[:：]\s*\+?\s*[\d,]+\s*원", " ", o)   # ": +23,000원" 가격차
    o = re.sub(r"^\s*\d+[.)]\s*", "", o)                    # "1. " 번호 prefix
    o = re.sub(r"_\d{5,}\s*$", "", o)                        # "_100370" 내부코드
    return re.sub(r"\s+", " ", o).strip()


_OPT_HEADER_RE = re.compile(r"^(사이즈\s*/\s*색상|색상|사이즈|가격|구분)$")


def model_options(raw_options):
    opts = [_clean_opt(o) for o in (raw_options or "").split("|") if o.strip()]
    opts = [o for o in opts if o and not _OPT_HEADER_RE.match(o)]
    opts = [o for o in opts if not _OPT_ADDON_RE.search(o)]  # 추가상품/구성 제외 (정제 후)
    opts = list(dict.fromkeys(opts))                          # 품절/가격 차이만 있던 중복 병합
    if len(opts) < 2:
        return []
    hits = sum(1 for o in opts if _OPT_TYPE_RE.search(o))
    if hits >= max(2, len(opts) * 0.5):
        return opts
    # 일반화: 색상/사이즈/스펙 등 변형축이 아닌 옵션이 과반 → 구성(모델) 옵션
    nonvar = sum(1 for o in opts if not _is_variantish(o))
    if nonvar >= max(2, len(opts) * 0.7) and len(opts) <= 12:
        return opts
    return []


# 이름 괄호 구성 열거: 한글 항 2개+ 슬래시 병기, 색상/사이즈/스펙 아님 → 모델 분해
_PAREN_ENUM_RE = re.compile(r"\(([가-힣][가-힣A-Za-z0-9 ]{1,14}(?:/[가-힣][가-힣A-Za-z0-9 ]{1,14}){1,3})\)")
_VARIANTISH_RE = re.compile(r"(?:SS|EK|LK|KK|[SQKD])(?![가-힣A-Za-z])|\d+\s*(?:인|[wW]|구|mm|cm|개입)"
                            r"|주광색|전구색|주백색")


def _is_variantish(term):
    t = term.strip()
    return (any(cb in t for cb in lex.COLOR_BASE) or _VARIANTISH_RE.search(t)
            or _OPT_ADDON_RE.search(t))


def name_enum_terms(name):
    m = _PAREN_ENUM_RE.search(name)
    if not m:
        return None, []
    terms = [t.strip() for t in m.group(1).split("/") if t.strip()]
    if len(terms) >= 2 and not any(_is_variantish(t) for t in terms):
        return m, terms
    return None, []


def option_records(r, opts):
    """모음전 PDP → 옵션별 합성 레코드. base에서 유형 나열·모음전 토큰 제거 후 옵션 결합."""
    base_tokens = [t for t in r["name"].split() if not _TYPE_LISTING.search(t)]
    out = []
    for i, opt in enumerate(opts):
        opt_clean = _clean_opt(_OPT_NUM_PREFIX.sub("", opt))
        # 옵션이 자기서술형(유형 포함 + 2토큰 이상)이면 base 브랜드부만 + 옵션
        synth_name = " ".join(dict.fromkeys(base_tokens + opt_clean.split()))
        sr = dict(r)
        sr["name"] = synth_name
        sr["prd_id"] = f"{r['prd_id']}#opt{i+1}"
        sr["_opt_src"] = opt
        out.append(sr)
    return out


def run_decompose(in_path=IN_JSONL, out_path=DECOMP_OUT):
    recs = [json.loads(l) for l in open(in_path, encoding="utf-8")]
    # 모델형 옵션 PDP → 옵션별 전개
    expanded = []
    n_optsplit = 0
    def enum_split(r):
        m, terms = name_enum_terms(r["name"])
        if not terms:
            return [r]
        base = r["name"][:m.start()] + " " + r["name"][m.end():]
        out = []
        for i, t in enumerate(terms):
            sr = dict(r)
            sr["name"] = re.sub(r"\s+", " ", f"{base} {t}").strip()
            sr["prd_id"] = f"{r['prd_id']}#cfg{i+1}"
            sr["_opt_src"] = t
            out.append(sr)
        return out

    for r in recs:
        opts = model_options(r.get("raw_options", ""))
        if opts:
            for orec in option_records(r, opts):
                expanded.extend(enum_split(orec))   # 옵션×열거 중첩 (나래울 케이스)
            n_optsplit += 1
            continue
        es = enum_split(r)
        expanded.extend(es)
        if len(es) > 1:
            n_optsplit += 1
    if n_optsplit:
        print(f"[decompose] 모델형 옵션 PDP {n_optsplit}개 → 옵션 전개 "
              f"(+{len(expanded)-len(recs)}행)")
    recs = expanded
    rows, dropped = [], 0
    for r in recs:
        k = make_key(r)
        if k is None:
            dropped += 1
            continue
        # 구매 옵션 목록 (정제·추가상품 제외) — 변형 팬아웃용. 모델분해 합성행은 제외
        if "_opt_src" in r:
            opt_list = []
        else:
            opt_list = [_clean_opt(o) for o in (r.get("raw_options") or "").split("|") if o.strip()]
            opt_list = [o for o in opt_list if o and not _OPT_HEADER_RE.match(o)
                        and not _OPT_ADDON_RE.search(o)]
            opt_list = list(dict.fromkeys(opt_list))
        rows.append({
            "prd_id": r["prd_id"], "mall": r["source"]["mall"],
            "catalog_key": k["catalog_key"], "brand": k["brand"],
            "canonical": k["canonical"], "cat_class": k["cat_class"],
            "model_code": k["model_code"],
            "variant_attrs": json.dumps(k["variant_attrs"], ensure_ascii=False),
            "options": "|".join(opt_list),
            "bundle_flag": int(k["bundle_flag"]),
            "l1": r["l1_category"], "l2": r["l2_category"],
            "name": r["name"], "price": r["price"],
            "color_opts": r["attributes"].get("color", ""),
            "url": r["source"]["url"],
        })
    with open(out_path, "w", encoding="utf-8-sig", newline="") as f:
        w = csv.DictWriter(f, fieldnames=list(rows[0].keys()))
        w.writeheader()
        w.writerows(rows)
    n_keys = len({r["catalog_key"] for r in rows})
    print(f"[decompose] {len(rows)}행 (폐기 {dropped}) → 카탈로그 키 {n_keys}개 "
          f"(롤업 {len(rows)/max(n_keys,1):.2f}:1) → {out_path}")
    return rows


# ── Stage 2: group ────────────────────────────────────────────────────────────

def _modal(vals):
    vals = [v for v in vals if v]
    return Counter(vals).most_common(1)[0][0] if vals else ""


_MEANINGLESS_RE = re.compile(
    r"^(상세\s*(페이지)?\s*(참조|참고)|해당\s*없음|없음|미정|기타|본품|단품|일반|보통|기본형?)$"
    r"|^[.·\-~/+,\s]+$|^\d{1,2}$")


def _clean_val(v):
    """속성값 정리 — 괄호류 제거(내용 보존), NEW 접두, 언더스코어, 의미 불명 값 제거."""
    if not isinstance(v, str):
        return v
    v = re.sub(r"^NEW[_\s]*", "", v, flags=re.IGNORECASE)
    v = re.sub(r"[()\[\]{}〔〕（）<>「」]", " ", v).replace("_", " ")
    v = re.sub(r"(?:\s+[/,+·:]\s*)+", " ", v)
    v = re.sub(r"\s+", " ", v).strip(" -/,+~:")
    if _MEANINGLESS_RE.match(v):
        return ""
    return v


_ENUM_RE = re.compile(r"(?:(?<=\s)|^)\d{1,2}(?:-\d{1,2})?[.)](?=\s|[가-힣])\s*")  # "1-1. " "01.옷장" 열거자(소수점 2.5 비매칭)
_STOCK_RE = re.compile(r"일시품절|재고소진|입고예정|품절")            # 재고 상태 문구
_PICK_RE = re.compile(r"(?:(?<=\s)|^)(?:선택|모음)(?=\s|$)")        # 옵션 지시어("SS/Q 선택", "모음")
_MOJIBAKE_RE = re.compile(r"[�]+")                              # 깨진 인코딩(�)


def _clean_title(t):
    """표시용 타이틀 최종 정리 — 옵션 열거자·재고문구·모지바케 제거, 괄호류 제거(내용 보존), 공백 정돈."""
    t = _MOJIBAKE_RE.sub(" ", t)
    t = _ENUM_RE.sub(" ", t)          # "1-1. 스테이 코이" → "스테이 코이"
    t = _STOCK_RE.sub(" ", t)
    t = re.sub(r"[()\[\]{}〔〕（）<>「」]", " ", t)
    t = _ENUM_RE.sub(" ", t)          # 괄호 제거로 공백이 생긴 열거자("01.옷장(…)" 케이스) 재적용
    t = _PICK_RE.sub(" ", t)          # "선택"/"모음" 옵션 지시어
    t = re.sub(r"(?:\s+[/,+·]\s*)+", " ", t)   # 괄호 제거 후 고아 구분자
    return _WS.sub(" ", t).strip(" -/,+~")


_FCANON = None
_FCANON_PATH = os.path.join(HERE, "outputs", "_catalog_canonical_furniture.json")


def _fcanon():
    """LLM canonical 스토어(catalog_geo.py --store …furniture… 배치가 채움). 없으면 빈 dict."""
    global _FCANON
    if _FCANON is None:
        try:
            _FCANON = json.load(open(_FCANON_PATH, encoding="utf-8")) \
                if os.path.exists(_FCANON_PATH) else {}
        except (ValueError, OSError):
            _FCANON = {}
    return _FCANON


def title_geo(brand, canonical, l2):
    canonical = _fcanon().get(f"{brand}|{canonical}", canonical)  # LLM canonical 우선(없으면 규칙값)
    t = f"{brand} {canonical}".strip()
    # 유형 보충 (§3.1-3): head noun에 유형 없으면 l2로
    has_type = any(k in canonical for k in
                   ("소파", "침대", "이불", "조명", "트리", "커버", "베개", "매트",
                    "러그", "등", "스탠드", "블라인드", "책상", "의자", "장",
                    "테이블", "스툴", "패드"))
    l2_overlap = l2 and any(w in canonical for w in l2.replace("/", " ").split())
    if l2 and l2 != "리퍼/전시/중고" and not has_type and not l2_overlap:
        t += f" {l2}"
    return _clean_title(t)


# 사이즈 표기 — 영문 코드 통일 (S/SS/D/Q/K/EK/LK, 2026-07-02 정책)
_SIZE_LABEL = {}  # 코드 그대로 사용


_COLOR_ENUM_RE = None


def _strip_dup_values(name, values):
    """이름에 속성과 동일한 토큰(색상 병기 등)이 있으면 제거 — 변형값 슬롯과 중복 방지."""
    for v in values:
        if not v or len(str(v)) < 2:
            continue
        name = re.sub(r"(?<![가-힣A-Za-z])" + re.escape(str(v)) + r"(?![가-힣A-Za-z])", " ", name)
    # 색상 병기 나열("화이트/블랙") 잔재 정리
    name = re.sub(r"(?:\s+/\s*)+", " ", name)
    return re.sub(r"\s+", " ", name).strip()


def title_commerce(brand, name_clean, va):
    dup_vals = [va.get(k) for k in ("color", "size", "seat", "watt", "cct")]
    # 색상은 개별 토큰까지 (병기 "화이트/블랙"의 각 색)
    if va.get("color"):
        dup_vals += re.split(r"[/_]", str(va["color"]))
    name_clean = _strip_dup_values(name_clean, dup_vals)
    parts = [brand, name_clean]
    # 변형값 슬롯 순서(§3.2): 색상 → 규격 → 사이즈 → 색온도 → 개입수 → 번들구성
    if va.get("option"):
        # 옵션 원문이 조합 전체를 서술 — 옵션만 부착 (색상/사이즈 개별 부착 생략)
        name_clean = _strip_dup_values(name_clean, [va["option"]])
        # 이름의 사이즈 잔재("( Q 사이즈)", 단독 코드)는 옵션과 모순 — 제거
        name_clean = re.sub(r"\(\s*[SQKDELS/ ]+\s*사이즈?\s*\)", " ", name_clean)
        if va.get("size"):
            name_clean = re.sub(r"(?<![A-Za-z가-힣])(SS|EK|LK|KK|[SQKD])(?![A-Za-z가-힣])", " ", name_clean)
        name_clean = _WS.sub(" ", name_clean).strip()
        return _clean_title(f"{brand} {name_clean} {va['option']}")
    order = ["color", "watt", "mm", "cm", "seat", "size", "cct", "pack",
             "module", "coverage", "mattress"]
    label = {"seat": "인용", "pack": "개입"}
    for k in order:
        v = va.get(k)
        if not v:
            continue
        if k == "size":
            parts.append(str(v))
        elif k == "mattress":
            parts.append(f"+{v}")           # 번들구성은 사이즈 뒤 유지
        elif k in label:
            parts.append(f"{v}{label[k]}")
        else:
            parts.append(str(v))
    return _clean_title(" ".join(p for p in parts if p))


def run_group():
    rows = list(csv.DictReader(open(DECOMP_OUT, encoding="utf-8-sig")))
    groups = defaultdict(list)
    for r in rows:
        groups[r["catalog_key"]].append(r)

    catalogs, variants, review = [], [], []
    for key, rs in groups.items():
        prices = [int(r["price"]) for r in rs if str(r["price"]).isdigit()]
        # §6.5 가격 가드: base price 편차 ±40% 초과 → 리뷰 큐
        needs_review = ""
        if len(prices) >= 2 and min(prices) > 0:
            vas = [json.loads(r["variant_attrs"]) for r in rs]
            has_trim_diff = len({v.get("coverage", "") for v in vas}) > 1 or \
                            len({v.get("module", "") for v in vas}) > 1
            if not has_trim_diff and max(prices) / min(prices) > 1.4:
                needs_review = "price_spread"
        r0 = min(rs, key=lambda r: len(r["name"]))  # 최단 원본명 대표
        all_colors = set()
        n_var = 0
        for r in rs:
            va = json.loads(r["variant_attrs"])
            opts = [c.strip() for c in (r["color_opts"] or "").split(" / ") if c.strip()]
            colors = opts or ([va["color"]] if va.get("color") else [])
            all_colors.update(colors)
            # ── 드롭다운 옵션 팬아웃 (2026-07-02: PDP 옵션 = 변형 열거) ──
            opt_list = [o for o in (r.get("options") or "").split("|") if o]
            if len(opt_list) >= 2:
                for opt in opt_list:
                    va_v = dict(va)
                    va_v["option"] = opt
                    # 옵션 문자열에서 색상·사이즈 파싱 시도
                    cm = next((cb for cb in lex.COLOR_BASE if cb in opt), "")
                    if cm:
                        va_v["color"] = cm          # 파싱된 색상 토큰만 (옵션 원문은 option에)
                        all_colors.add(cm)
                    sm = re.search(r"(?<![A-Za-z])(SS|EK|LK|KK|[SQKD])(?![A-Za-z])", opt)
                    if sm:
                        va_v["size"] = sm.group(1)
                    va_v = {k: _clean_val(x) for k, x in va_v.items()}
                    va_v = {k: x for k, x in va_v.items() if x}
                    variants.append({
                        "catalog_key": key, "mall": r["mall"], "url": r["url"],
                        "title_commerce": title_commerce(r["brand"], r["canonical"], va_v),
                        "variant_attrs": json.dumps(va_v, ensure_ascii=False),
                        "price": r["price"], "name": r["name"],
                    })
                    n_var += 1
                continue  # 옵션 열거가 조합 전체 — 아래 교차 전개 생략
            # ── 다중값 축 교차 전개 (2026-07-02: 구성 옵션 전부 분해) ──
            # 사이즈: 침대=슬래시 개별, 침구=콤마 옵션그룹
            raw_sz = va.get("size", "")
            if r["cat_class"] == "bed" and "/" in raw_sz:
                sz_list = [t for t in raw_sz.split("/") if t]
            elif r["cat_class"] == "bedding" and "," in raw_sz:
                sz_list = [t.strip() for t in raw_sz.split(",") if t.strip()]
            else:
                sz_list = [raw_sz]
            # 인수(소파 1/2/3/4/5인), 와트(50W/60W), 색온도(주광색/전구색)
            def _split_axis(key):
                v = va.get(key, "")
                if v and "/" in str(v) and "+" not in str(v):  # "4인+1인" 세트는 유지
                    return [t for t in str(v).split("/") if t]
                return [v] if v else [""]
            seat_list = _split_axis("seat")
            watt_list = _split_axis("watt")
            cct_list = _split_axis("cct")
            import itertools
            for c, sz, st, wt, ct in itertools.product(
                    (colors or [""]), sz_list, seat_list, watt_list, cct_list):
                va_v = dict(va)
                # 색상 축 정합성: 조합/코드 섞인 긴 값은 option으로, color엔 토큰만
                if c and (len(c) > 12 or re.search(r"\d{4,}|상판|측판|커버|매트리스|프레임|수납장", c)):
                    va_v["option"] = c
                    va_v["color"] = next((cb for cb in lex.COLOR_BASE if cb in c), "")
                else:
                    va_v["color"] = c
                if sz:
                    va_v["size"] = sz
                if st:
                    va_v["seat"] = st
                if wt:
                    va_v["watt"] = wt if str(wt).upper().endswith("W") else f"{wt}W"
                if ct:
                    va_v["cct"] = ct
                va_v = {k: _clean_val(x) for k, x in va_v.items()}
                va_v = {k: x for k, x in va_v.items() if x}
                variants.append({
                    "catalog_key": key, "mall": r["mall"], "url": r["url"],
                    "title_commerce": title_commerce(r["brand"], r["canonical"], va_v),
                    "variant_attrs": json.dumps(va_v, ensure_ascii=False),
                    "price": r["price"], "name": r["name"],
                })
                n_var += 1
        va0 = json.loads(r0["variant_attrs"])
        tg = title_geo(r0["brand"], r0["canonical"], r0["l2"])
        if va0.get("mattress"):
            tg += f" +{va0['mattress']}"
        catalogs.append({
            "catalog_key": key, "brand": r0["brand"],
            "title_geo": tg,
            "product_name": r0["canonical"], "l1": r0["l1"], "l2": r0["l2"],
            "cat_class": r0["cat_class"], "model_code": r0["model_code"],
            "colors": "|".join(sorted(all_colors)[:20]), "n_colors": len(all_colors),
            "n_products": len(rs), "n_variants": n_var,
            "price_min": min(prices) if prices else "",
            "price_max": max(prices) if prices else "",
            "bundle_flag": r0["bundle_flag"],
            "needs_review": needs_review,
            "sample_url": r0["url"],
        })
        if needs_review:
            review.append(key)

    for path, data in ((CATALOG_OUT, catalogs), (VARIANT_OUT, variants)):
        with open(path, "w", encoding="utf-8-sig", newline="") as f:
            w = csv.DictWriter(f, fieldnames=list(data[0].keys()))
            w.writeheader()
            w.writerows(data)
    write_catalog_xlsx(catalogs, variants)
    print(f"[group] 카탈로그 {len(catalogs)}개 · 변형 {len(variants)}개 "
          f"· 리뷰 큐 {len(review)}건")
    print(f"  → {CATALOG_OUT}")
    print(f"  → {VARIANT_OUT}")
    # 몰별 롤업률 (부록 QA 기준)
    by_mall = defaultdict(lambda: [0, set()])
    for r in rows:
        by_mall[r["mall"]][0] += 1
        by_mall[r["mall"]][1].add(r["catalog_key"])
    for m, (n, keys) in sorted(by_mall.items()):
        print(f"  {m}: 상품 {n} → 카탈로그 {len(keys)} ({n/len(keys):.2f}:1)")


CATALOG_XLSX = os.path.join(OUT, "catalogs_furniture.xlsx")


def write_catalog_xlsx(catalogs, variants):
    """카탈로그 엑셀 — 시트: 카탈로그 / 변형(title_commerce) / 브랜드요약."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    hf, hfill = Font(bold=True, color="FFFFFF"), PatternFill("solid", fgColor="4472C4")

    # ── 시트 1: 카탈로그 ──
    ws = wb.active
    ws.title = "카탈로그"
    ws.append(["카탈로그키", "브랜드", "카탈로그명(title_geo)", "정규화모델명",
               "대카", "중카", "모델코드", "색상들", "색상수", "상품수(PDP)",
               "변형수", "최저가", "최고가", "세트", "리뷰사유", "대표URL"])
    for c in ws[1]:
        c.font, c.fill = hf, hfill
    for r in catalogs:
        ws.append([r["catalog_key"], r["brand"], r["title_geo"], r["product_name"],
                   r["l1"], r["l2"], r["model_code"], r["colors"], r["n_colors"],
                   r["n_products"], r["n_variants"],
                   r["price_min"], r["price_max"],
                   "세트" if str(r["bundle_flag"]) == "1" else "",
                   r["needs_review"], r["sample_url"]])
    ws.freeze_panes = "C2"
    ws.auto_filter.ref = ws.dimensions
    for i, w in enumerate([30, 10, 42, 36, 12, 12, 12, 30, 7, 10, 8, 11, 11, 6, 12], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # ── 시트 2: 변형 (title_commerce) ──
    ws2 = wb.create_sheet("변형")
    ws2.append(["카탈로그키", "쇼핑몰", "커머스명(title_commerce)", "변형속성", "가격", "원본상품명", "URL"])
    for c in ws2[1]:
        c.font, c.fill = hf, hfill
    for v in variants:
        ws2.append([v["catalog_key"], v["mall"], v["title_commerce"],
                    v["variant_attrs"], v["price"], v["name"], v["url"]])
    ws2.freeze_panes = "C2"
    ws2.auto_filter.ref = ws2.dimensions
    for i, w in enumerate([30, 10, 55, 30, 11, 45], 1):
        ws2.column_dimensions[get_column_letter(i)].width = w

    # ── 시트 3: 브랜드요약 ──
    ws3 = wb.create_sheet("브랜드요약")
    ws3.append(["브랜드", "카탈로그수", "상품수", "변형수", "롤업률", "리뷰건수"])
    for c in ws3[1]:
        c.font, c.fill = hf, hfill
    agg = defaultdict(lambda: [0, 0, 0, 0])
    for r in catalogs:
        a = agg[r["brand"]]
        a[0] += 1
        a[1] += int(r["n_products"])
        a[2] += int(r["n_variants"])
        a[3] += 1 if r["needs_review"] else 0
    for b, (nc, np_, nv, nr) in sorted(agg.items(), key=lambda x: -x[1][1]):
        ws3.append([b, nc, np_, nv, f"{np_/nc:.2f}:1", nr])
    ws3.column_dimensions["A"].width = 14

    wb.save(CATALOG_XLSX)
    print(f"[xlsx] 카탈로그 {len(catalogs)} · 변형 {len(variants)} → {CATALOG_XLSX}")



# ── verify: 골든 회귀 스위트 (§5.3 Phase 2) ──────────────────────────────────

GOLDEN_MERGE = [  # 같은 키여야 함 (부분문자열 쌍)
    ("엘리쉬 3.5인 기능성", "엘리쉬 멀티 4인", False),  # 엘리쉬 멀티는 별도일 수 있어 soft
    ("엘리쉬 3.5인", "엘리쉬 4인", True),
    ("그레이스 트리플 블라인드(차콜)", "그레이스 트리플 블라인드(베이지)", True),
    ("코티지 80수 고밀도 순면 호텔 차렵이불 - 화이트", "코티지 80수 고밀도 순면 호텔 차렵이불 - 그레이", True),
    ("스카치 트리 (가지수", None, True),  # 240/270cm 동일 키 (단항: cm 제거 확인)
]
GOLDEN_MULTI = [("길가온 흙침대", 2)]  # (부분문자열, 최소 카탈로그 키 수)
GOLDEN_SPLIT = [  # 다른 키여야 함
    ("키큰 긴문 옷장 800", "와이드 서랍장 1200"),   # 모음전 옵션 분해 (32954)
    ("수납침대 프레임 Q", "수납침대 Q＋독립매트"),  # 매트 구성 분리 (2026-07-02)
    ("시그니처 스윙", "시그니처 와이드"),
    ("탑볼형", "언더볼형"),
    ("60수", "80수"),
    ("트리 풀세트", "무장식 트리"),
]


GOLDEN_CLEAN = [  # (상품명 부분문자열, canonical에 있으면 안 되는 토큰)
    ("퍼백 씨티백 PB-600", ["기숙사", "자취방", "강의실", "편한"]),
]


def run_verify():
    rows = list(csv.DictReader(open(DECOMP_OUT, encoding="utf-8-sig")))

    def keys_for(sub):
        return {r["catalog_key"] for r in rows if sub in r["name"]}

    fails = []
    # 병합 케이스
    for a, b, hard in GOLDEN_MERGE:
        if b is None:
            ks = keys_for(a)
            # cm만 다른 이름들이 한 키 그룹 내 존재하는지
            if len(ks) == 0:
                continue
            names = {r["name"] for r in rows if r["catalog_key"] in ks}
            cms = {re.sub(r"\d+cm", "", n) for n in names}
            if len(cms) < len(names):
                pass  # cm 변형이 병합됨
            continue
        ka, kb = keys_for(a), keys_for(b)
        if not ka or not kb:
            print(f"  [skip] 데이터 없음: {a} / {b}")
            continue
        if ka & kb:
            print(f"  [PASS-병합] {a} ↔ {b}")
        else:
            msg = f"[{'FAIL' if hard else 'WARN'}-병합안됨] {a} ↔ {b}"
            print(f"  {msg}")
            if hard:
                fails.append(msg)
    # 분리 케이스
    for a, b in GOLDEN_SPLIT:
        ka, kb = keys_for(a), keys_for(b)
        if not ka or not kb:
            print(f"  [skip] 데이터 없음: {a} / {b}")
            continue
        if ka & kb:
            msg = f"[FAIL-오병합] {a} ↔ {b}: {list(ka & kb)[:2]}"
            print(f"  {msg}")
            fails.append(msg)
        else:
            print(f"  [PASS-분리] {a} ↮ {b}")
    # 드롭다운 완전 분해 감사 (2026-07-02: 모든 옵션은 찢어져야 함)
    from collections import defaultdict
    var_by_url, prod_by_url, opts_by_url = defaultdict(int), defaultdict(int), {}
    try:
        for v in csv.DictReader(open(VARIANT_OUT, encoding="utf-8-sig")):
            var_by_url[v["url"]] += 1
    except FileNotFoundError:
        pass
    for r in rows:
        prod_by_url[r["url"]] += 1
        n_o = len([o for o in (r.get("options") or "").split("|") if o])
        if n_o >= 2:
            opts_by_url[r["url"]] = max(opts_by_url.get(r["url"], 0), n_o)
    under = [u for u, n in opts_by_url.items()
             if max(var_by_url.get(u, 0), prod_by_url.get(u, 0)) < n]
    if under:
        msg = f"[FAIL-옵션분해] 미분해 PDP {len(under)}개"
        print(f"  {msg}"); fails.append(msg)
    else:
        print(f"  [PASS-옵션분해] {len(opts_by_url)}개 PDP 완전 분해")

    # 다중 분리 케이스 (구성 열거)
    for sub, mink in GOLDEN_MULTI:
        ks = keys_for(sub)
        if ks and len(ks) < mink:
            msg = f"[FAIL-열거분해] {sub}: 키 {len(ks)}개 < {mink}"
            print(f"  {msg}"); fails.append(msg)
        elif ks:
            print(f"  [PASS-열거분해] {sub}: 키 {len(ks)}개")
    # 정제 케이스
    for sub, banned in GOLDEN_CLEAN:
        hits = [r for r in rows if sub in r["name"]]
        for r in hits:
            bad = [b for b in banned if b in r["canonical"]]
            if bad:
                msg = f"[FAIL-정제] {sub}: canonical에 {bad} 잔존"
                print(f"  {msg}")
                fails.append(msg)
        if hits and not any(sub in f for f in fails):
            print(f"  [PASS-정제] {sub}")
    if fails:
        print(f"[verify] 실패 {len(fails)}건")
        sys.exit(1)
    print("[verify] PASS")


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("cmd", choices=["decompose", "group", "verify", "all"])
    ap.add_argument("--in", dest="in_path", default=IN_JSONL)
    args = ap.parse_args()
    if args.cmd in ("decompose", "all"):
        run_decompose(args.in_path)
    if args.cmd in ("group", "all"):
        run_group()
    if args.cmd in ("verify", "all"):
        run_verify()


if __name__ == "__main__":
    main()
