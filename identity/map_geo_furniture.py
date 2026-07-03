#!/usr/bin/env python3
"""가구/인테리어 추출 CSV → 11번가 GEO 속성 스키마 매핑.

  python3 map_geo_furniture.py                 # furniture_all_brands.csv 전체
  python3 map_geo_furniture.py --in <csv>      # 입력 지정
  python3 map_geo_furniture.py --report-only   # 리포트만 재생성

입력: outputs/furniture_all_brands.csv (extract_all_furniture.py 산출)
출력:
  outputs/furniture_geo_mapped.jsonl  — 상품별 GEO 속성 레코드
  outputs/furniture_geo_report.md     — 대카별 속성 채움률 리포트

스키마 기준: ~/Work/bm-geo-attributes/가구인테리어/categories/*.md (v2.10)
  · 메타 공통 20 속성 키 사용 (product_type, manufacturer, color, size, material,
    width_cm/depth_cm/height_cm, count_pcs, country_of_origin, is_assembly_required,
    installation_service, safety_cert, warranty_period, model_no, brand_line …)
  · 대카/중카 분류: 스키마 md의 중카 목록 기반 키워드 매칭
  · 대카 델타 파생: lighting_type/light_source(조명), bedding_type/filling_type/
    bedding_size(침구), sofa_seat_count(거실가구), interior_item_type(인테리어소품)
정책: CP-04 근거 없으면 null(빈값) — 추론 금지, 원문 그대로.
"""
import argparse
import csv
import json
import os
import re
import sys

HERE = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, HERE)
import catalog_lexicon_furniture as lex  # noqa: E402

OUT_DIR = os.path.join(HERE, "outputs")
IN_DEFAULT = os.path.join(OUT_DIR, "furniture_all_brands.csv")
JSONL_OUT = os.path.join(OUT_DIR, "furniture_geo_mapped.jsonl")
REPORT_OUT = os.path.join(OUT_DIR, "furniture_geo_report.md")
XLSX_OUT = os.path.join(OUT_DIR, "furniture_geo_mapped.xlsx")

# ── 대카/중카 분류 규칙 ────────────────────────────────────────────────────────
# (대카, 중카, [키워드…]) — 긴 키워드 우선 매칭. 스키마 md 중카 목록 기준.
L2_RULES = [
    # 조명 (전구/조명)
    ("조명", "전구", ["led램프", "led전구", "전구", "램프", "형광등", "할로겐"]),
    ("조명", "조명", ["방등", "거실등", "주방등", "욕실등", "현관등", "센서등", "직부등",
                    "투광등", "평판등", "십자등", "일자등", "펜던트", "샹들리에",
                    "스탠드", "벽등", "무드등", "다운라이트", "레일조명", "라인조명",
                    "천장등", "등기구", "파티라이트", "스트링라이트", "앵두전구",
                    "조명"]),
    # 비침구 커버 (침구 규칙보다 먼저 — 카테고리 텍스트 오염 방지)
    ("홈패브릭/수예", "커버류", ["쇼파커버", "소파커버", "의자커버"]),
    # 침구 — 커버류가 본품보다 먼저 (베개커버가 패드로 새는 것 방지)
    ("침구", "베개커버", ["베개커버", "베개덮개", "베갯잇"]),
    ("침구", "침대/매트커버", ["매트리스커버", "침대커버", "이불커버"]),
    ("침구", "이불", ["차렵이불", "누비이불", "구스이불", "오리털이불", "이불솜", "이불"]),
    ("침구", "요/토퍼", ["토퍼", "요매트"]),
    ("침구", "패드/스프레드", ["침대패드", "패드", "스프레드"]),
    ("침구", "베개", ["베개", "경추베개", "메모리폼베개"]),
    ("침구", "침구세트", ["침구세트", "이불세트"]),
    ("침구", "유아동침구", ["유아이불", "아기이불", "낮잠이불"]),
    ("침구", "솜", ["충전솜", "솜"]),
    # 거실가구
    ("거실가구", "소파", ["리클라이너", "소파베드", "소파"]),
    ("거실가구", "거실장/TV장", ["거실장", "tv장", "tv 장", "티비장"]),
    ("거실가구", "진열장/장식장", ["진열장", "장식장"]),
    ("거실가구", "테이블", ["사이드테이블", "커피테이블", "좌탁", "테이블"]),
    # 침실가구
    ("침실가구", "침대", ["침대프레임", "벙커침대", "이층침대", "침대"]),
    ("침실가구", "매트리스", ["매트리스"]),
    ("침실가구", "협탁", ["협탁"]),
    ("침실가구", "서랍장", ["서랍장"]),
    ("침실가구", "장롱/붙박이장", ["장롱", "붙박이장", "옷장"]),
    ("침실가구", "화장대", ["화장대"]),
    ("침실가구", "거울", ["전신거울", "거울"]),
    # 주방가구
    ("주방가구", "식탁의자", ["식탁의자"]),
    ("주방가구", "식탁", ["식탁"]),
    ("주방가구", "레인지대", ["레인지대"]),
    ("주방가구", "왜건/카트", ["왜건", "카트"]),
    ("주방가구", "주방수납장", ["주방수납장", "그릇장"]),
    # 학생/사무용가구
    ("학생/사무용가구", "책상", ["컴퓨터책상", "책상"]),
    ("학생/사무용가구", "의자", ["게이밍의자", "사무의자", "메쉬의자", "학생의자", "의자"]),
    ("학생/사무용가구", "책장", ["책장"]),
    ("학생/사무용가구", "책꽂이", ["책꽂이"]),
    # 수납가구
    ("수납가구", "신발장", ["신발장"]),
    ("수납가구", "수납장", ["수납장", "행거", "리빙박스"]),
    ("수납가구", "수납선반", ["선반"]),
    # 커튼/블라인드 (소품보다 먼저 — 키워드 우선순위)
    ("커튼/블라인드", "커튼", ["암막커튼", "쉬어커튼", "커튼"]),
    ("커튼/블라인드", "블라인드", ["블라인드"]),
    ("커튼/블라인드", "롤스크린", ["롤스크린"]),
    # 카페트/러그 (소품보다 먼저 — "오브제 러그" 오분류 방지)
    ("카페트/러그", "러그", ["러그"]),
    ("카페트/러그", "카페트", ["카페트", "카펫"]),
    ("카페트/러그", "매트", ["발매트", "매트"]),
    # 인테리어소품
    ("인테리어소품", "크리스마스 트리", ["크리스마스", "트리", "오너먼트", "산타", "가랜드",
                                  "리스", "츄리"]),
    ("인테리어소품", "디퓨저", ["디퓨저"]),
    ("인테리어소품", "캔들", ["캔들", "향초"]),
    ("인테리어소품", "시계", ["벽시계", "시계"]),
    ("인테리어소품", "액자", ["액자", "포스터"]),
    ("인테리어소품", "기타 인테리어소품", ["오브제", "화병", "꽃병", "트레이", "모빌", "장식"]),
    # 홈패브릭/수예
    ("홈패브릭/수예", "쿠션", ["쿠션"]),
    ("홈패브릭/수예", "방석", ["방석"]),
    ("홈패브릭/수예", "커버류", ["쿠션커버"]),
    ("홈패브릭/수예", "주방패브릭", ["테이블보", "식탁보", "앞치마"]),
    # 아웃도어가구
    ("아웃도어가구", "라탄가구", ["라탄"]),
    ("아웃도어가구", "벤치", ["벤치"]),
    ("아웃도어가구", "야외용 테이블", ["야외테이블", "캠핑테이블"]),
    ("아웃도어가구", "야외용의자", ["야외의자", "캠핑의자"]),
]

# 브랜드 기본 대카 (이름 매칭 실패 시 폴백)
BRAND_DEFAULT_L1 = {
    "wooree": "조명", "bflamp": "조명", "vittz": "조명",
    "flora": "침구", "prielle": "침구",
    "jakomo": "거실가구",
    "dotoro": "인테리어소품",
    "dongsuh": None,       # 종합가구 — 이름 기반만
    "mothershome": None,   # 생활용품 — 이름 기반만
}

# ── 대카 델타 파생 규칙 ────────────────────────────────────────────────────────
LIGHTING_TYPE = [("천장등", ["방등", "거실등", "주방등", "천장등", "평판등", "십자등",
                          "일자등", "직부등", "등기구"]),
                 ("펜던트", ["펜던트"]), ("샹들리에", ["샹들리에"]),
                 ("스탠드", ["스탠드"]), ("벽등", ["벽등", "브라켓"]),
                 ("무드등", ["무드등", "수면등", "취침등"]),
                 ("다운라이트", ["다운라이트", "매입등"]),
                 ("라인조명", ["라인조명", "레일조명", "t5"]),
                 ("캠핑랜턴", ["랜턴"])]
LIGHT_SOURCE = [("LED", ["led"]), ("형광등", ["형광"]), ("백열전구", ["백열"]),
                ("할로겐", ["할로겐"])]
BEDDING_TYPE = [("베개커버", ["베개커버", "베개덮개", "베갯잇"]),
                ("매트리스커버", ["매트리스커버"]),
                ("침대커버", ["침대커버", "이불커버"]),  # 이불커버→침대커버 (스키마 허용값)
                ("토퍼", ["토퍼"]), ("패드", ["패드"]),
                ("베개", ["베개"]),
                ("차렵이불", ["차렵"]), ("누비이불", ["누비이불"]),
                ("침구세트", ["침구세트", "이불세트"]),
                ("이불", ["이불"])]
# 충전재가 없는 제품(커버·시트류) — filling_type 부여 금지
_NO_FILLING = {"베개커버", "매트리스커버", "침대커버", ""}
FILLING_TYPE = [("다운(거위털·오리털)", ["구스", "거위털", "오리털", "덕다운"]),
                ("마이크로화이버", ["마이크로화이버", "극세사솜"]),
                ("모달솜", ["모달솜"]), ("메모리폼", ["메모리폼"]),
                ("라텍스", ["라텍스"]), ("폴리에스터솜", ["폴리에스터", "폴리솜"])]
INTERIOR_ITEM = [("액자·포스터", ["액자", "포스터"]), ("거울", ["거울"]),
                 ("시계", ["시계"]), ("화병·꽃병", ["화병", "꽃병"]),
                 ("캔들·디퓨저", ["캔들", "디퓨저", "향초"]), ("트레이", ["트레이"]),
                 ("화분", ["화분"]), ("디자인 오브제", ["오브제"])]

# ── 확장 델타 추출기 (스키마 중카 카탈로그 속성 대응, 2026-07-02) ──────────────
# 조명: cat_41202 색온도 / cat_50108 소비전력 / cat_49270 소켓 / cat_49264 등수 / cat_43293 길이
_CCT_TOKEN_RE = re.compile(r"주광색|전구색|주백색")
_KELVIN_RE = re.compile(r"(\d{4})\s*[kK]\b")
_POWER_RE = re.compile(r"(\d{1,3}(?:\.\d)?)\s*[wW]\b")
_SOCKET_RE = re.compile(r"(?<!\d)(\d{1,3})\s*구\b")
_FIXTURE_RE = re.compile(r"(?<![가-힣])(\d{1,2})\s*등\b")
_LEN_MM_RE = re.compile(r"\(?(\d{3,4})\s*mm\)?")
_INCH_RE = re.compile(r"(\d{1,2})\s*인치")
_IP_RE = re.compile(r"IP(\d{2})\b")
_PYEONG_RE = re.compile(r"(\d{1,3})\s*평형?\b")


def light_color_temp(name, raw_color):
    """색온도: 상품명·옵션 원문의 주광/전구/주백색 토큰 (color에서 제외한 것을 여기로)."""
    toks = []
    for src in (name, raw_color):
        for t in _CCT_TOKEN_RE.findall(src or ""):
            if t not in toks:
                toks.append(t)
    if toks:
        return " / ".join(toks)
    m = _KELVIN_RE.search(name or "")
    return f"{m.group(1)}K" if m else ""


def _first(pattern, text, suffix=""):
    m = pattern.search(text or "")
    return f"{m.group(1)}{suffix}" if m else ""


# 소파: cat_48827 마감재 / cat_42446 형태 / cat_48438 구성
_FINISH_GRADES = ["천연면피", "통가죽", "전체가죽", "전면가죽", "전측면가죽",
                  "이태리", "나파", "슈렁큰", "아닐린"]
_SOFA_FORMS = [("코너카우치형", ["코너카우치"]), ("카우치형", ["카우치"]),
               ("코너형", ["코너"]), ("모듈형", ["모듈"]),
               ("리클라이너형", ["리클라이너"]), ("소파베드형", ["소파베드"])]
_COMPO_RE = re.compile(r"\+\s*[가-힣A-Za-z0-9 ]{1,15}(?:\([^)]*\))?")


def sofa_attrs(name):
    finish = " / ".join(g for g in _FINISH_GRADES if g in name)
    form = first_match(name, _SOFA_FORMS)
    compo = " ".join(c.strip() for c in _COMPO_RE.findall(name))[:80]
    return finish, form, compo


# 침구 subtype (스키마 이불 하위유형)
BEDDING_SUBTYPE = [("거위털이불", ["구스이불", "거위털이불", "구스 차렵", "블렌딩 구스"]),
                   ("오리털이불", ["오리털이불", "덕다운 이불"]),
                   ("차렵이불", ["차렵"]), ("홑이불", ["홑이불", "홑겹이불"]),
                   ("이불커버", ["이불커버"]), ("이불세트", ["이불세트"]),
                   ("담요", ["담요", "블랭킷"])]

# 트리 subtype (스키마 크리스마스 트리 하위유형)
# 주의: 매칭 전 "크리스마스" 토큰 제거 필수 ("~마스리스" 부분매칭 방지) — tree_subtype() 사용
TREE_SUBTYPE = [("무장식트리", ["무장식"]), ("미니트리", ["미니트리", "미니 트리"]),
                ("가렌드", ["가랜드", "가렌드"]),
                ("인형", ["인형"]), ("산타모자", ["산타모자"]), ("산타복", ["산타복", "산타 복"]),
                ("조명", ["파티라이트", "앵두전구", "스트링", "와이어전구", "커튼전구"]),
                # 액세서리 — 트리 본체 판정보다 먼저 (탑별/스커트가 일반트리로 새는 것 방지)
                ("장식용품", ["오너먼트", "장식볼", "장식세트", "리본", "지팡이", "선물상자",
                          "트리스커트", "트리 스커트", "스커트", "탑별", "볼세트", "픽", "베너",
                          "오르골", "스노우볼", "기차", "장식"]),
                ("리스", ["리스"]),
                ("일반트리", ["트리"])]


def tree_subtype(name):
    t = (name or "").replace("크리스마스", " ")  # "크리스마스"의 "리스" 오탐 차단
    return first_match(t, TREE_SUBTYPE)

# 특징 (cat_46459/cat_137965 자유값) — 근거 토큰만
_FEATURE_TOKENS = ["방수", "냉감", "항균", "알러지케어", "암막", "센서", "리모컨",
                   "무선", "충전식", "타이머", "밝기조절", "높이조절", "각도조절"]

# 향 (cat_50873 — 디퓨저/캔들)
_SCENTS = [("시트러스", ["시트러스", "레몬", "자몽", "오렌지향"]), ("플로럴", ["플로럴", "장미", "로즈향", "자스민"]),
           ("우디", ["우디", "샌달우드", "우드향"]), ("머스크", ["머스크"]),
           ("라벤더향", ["라벤더향"]), ("무향", ["무향"])]


def parse_features(name):
    out = []
    for t in _FEATURE_TOKENS:
        if t not in name:
            continue
        if t == "센서" and re.search(r"일반\s*/\s*센서", name):
            continue  # 일반/센서 선택옵션 — 전 제품 기능 아님
        out.append(t)
    return " / ".join(out)[:60]



def first_match(text, rules):
    low = text.lower()
    for label, kws in rules:
        if any(k in low for k in kws):
            return label
    return ""


# 스키마 밖 상품 (가구/인테리어 카테고리에 없음) → 미분류
_OUT_OF_SCHEMA = ["실링팬", "멀티탭", "콘센트", "우산", "양산"]


def classify(name, category, source):
    """(대카, 중카) 분류: 상품명+카테고리 키워드 → 브랜드 기본 대카 폴백."""
    text = f"{name} {category}".lower()
    if any(k in text for k in _OUT_OF_SCHEMA):
        return "", ""
    text = text.replace("플러그", "")  # "플러그"의 "러그" 오분류 방지
    for l1, l2, kws in L2_RULES:
        if any(k in text for k in kws):
            return l1, l2
    default = BRAND_DEFAULT_L1.get(source)
    if default:
        return default, ""
    return "", ""


# 수량: N개입/N매입/Np/Np세트/N팩/(장식:N개). "N구"(소켓수)·"N인"(좌석)은 제외.
_COUNT_RE = re.compile(
    r"(\d+)\s*(?:개입|개 입|매입|매 입|팩|[pP](?:cs)?(?=\s|세트|$|\))|입\b|매\b|개\s*세트|개\b(?![월년]))"
    r"|(\d)\s*\+\s*(\d)(?!\s*인)")
# 번들 사은품 구간 제거: "+세탁망 1P", "+쿠션2개" — 본품 수량 아님
_BUNDLE_RE = re.compile(r"\+\s*[가-힣A-Za-z ]{1,12}\d+\s*(?:개|매|[pP])\b(?:\s*\([^)]*\))?")


def parse_count(name):
    t = _BUNDLE_RE.sub(" ", name or "")
    m = _COUNT_RE.search(t)
    if not m:
        return ""
    if m.group(1):
        return m.group(1)
    return str(int(m.group(2)) + int(m.group(3)))  # 1+1 → 2


# ── 상품명 기반 보강 파서 (OCR 전 무료 채움) ──────────────────────────────────
# 단일 치수: "180cm 트리", "높이 150cm" — 트리/스탠드류는 높이로 해석
_SINGLE_DIM_RE = re.compile(r"(?<![\dxX×*.])(\d{1,3})\s*(?:cm|㎝)(?![\dxX×*])")
# 2차원: "150x200", "150×200cm" — 침구/러그 폭×길이
_TWO_DIM_RE = re.compile(r"(\d{2,3})\s*[xX×*]\s*(\d{2,3})\s*(?:cm|㎝)?")
# 3차원: "1200x600x750"
_THREE_DIM_RE = re.compile(r"(\d{2,4})\s*[xX×*]\s*(\d{2,4})\s*[xX×*]\s*(\d{2,4})\s*(mm|cm|㎝)?")


_REF_DIM_RE = re.compile(r"(?:추천|호환|적용|대응)[^)]{0,10}[:：]?\s*\d{2,3}\s*(?:cm|㎝)")


def name_height_cm(name):
    """상품명 단일 cm 표기 → 높이 (트리·스탠드·행거류)."""
    t = _REF_DIM_RE.sub(" ", name or "")  # "(추천트리:270cm)" 등 참조 치수 제거
    if _TWO_DIM_RE.search(t):
        return ""
    m = _SINGLE_DIM_RE.search(t)
    return m.group(1) if m else ""


def name_two_dim(name):
    """상품명 2차원 표기 → (가로, 세로) cm. 침구·러그·매트."""
    m = _TWO_DIM_RE.search(name or "")
    if not m:
        return "", ""
    w, d = int(m.group(1)), int(m.group(2))
    if 30 <= w <= 400 and 30 <= d <= 400:
        return str(w), str(d)
    return "", ""


def name_three_dim(name):
    """상품명 3차원 표기 → (W, D, H) cm. mm면 /10."""
    m = _THREE_DIM_RE.search(name or "")
    if not m:
        return "", "", ""
    w, d, h = float(m.group(1)), float(m.group(2)), float(m.group(3))
    unit = (m.group(4) or "").lower()
    if unit == "mm" or (not unit and max(w, d, h) > 400):
        w, d, h = w / 10, d / 10, h / 10
    return (f"{w:g}", f"{d:g}", f"{h:g}")


# 상품명 색상 토큰 (긴 것 우선 정렬)
_NAME_COLORS = sorted(
    ["로즈골드", "로즈", "화이트", "블랙", "그레이", "베이지", "브라운", "월넛", "오크", "네이비",
     "블루", "핑크", "옐로우", "그린", "레드", "아이보리", "크림", "카키",
     "올리브", "민트", "라벤더", "골드", "실버", "차콜", "와인",
     "우드", "내추럴", "체리", "메이플", "퍼플", "오렌지"],
    key=len, reverse=True)


def name_color(name):
    low = (name or "")
    found = []
    for c in _NAME_COLORS:
        if c in low and not any(c in f for f in found):  # 긴 토큰의 부분문자열 제거
            found.append(c)
    return " / ".join(found[:3])


# 상품명 소재 키워드 → GEO 표준 소재 (긴 것 우선)
_NAME_MATERIALS = [
    ("천연가죽", "가죽(천연)"), ("통가죽", "가죽(천연)"), ("소가죽", "가죽(천연)"),
    ("면피", "가죽(천연)"), ("나파가죽", "가죽(천연)"),
    ("인조가죽", "인조가죽(PU)"), ("합성가죽", "인조가죽(PU)"), ("레자", "인조가죽(PU)"),
    ("패브릭", "패브릭"), ("아쿠아텍스", "패브릭"), ("부클", "패브릭"),
    ("원목", "원목"), ("고무나무", "원목"), ("애쉬", "원목"), ("편백", "원목"),
    ("라탄", "라탄"), ("대리석", "대리석"), ("세라믹", "세라믹"),
    ("스테인리스", "스테인리스"), ("메탈", "메탈"), ("철제", "메탈"),
    ("유리", "글래스"), ("아크릴", "플라스틱"),
    ("메모리폼", "메모리폼"), ("라텍스", "라텍스"),
    ("구스", "다운"), ("거위털", "다운"), ("오리털", "다운"),
    ("극세사", "폴리에스터"), ("마이크로화이버", "폴리에스터"),
    ("순면", "면(코튼)"), ("코튼", "면(코튼)"), ("리넨", "면(코튼)"), ("린넨", "면(코튼)"),
]


def name_material(name):
    low = (name or "")
    for kw, std in _NAME_MATERIALS:
        if kw in low:
            return std
    return ""


_SEAT_COMBO_RE = re.compile(r"(?<![\d.])([1-6](?:\.5)?)\s*인\s*\+\s*([1-6](?:\.5)?)\s*인")
_SEAT_RE = re.compile(r"(?<![\d.])([1-6](?:\.5)?)\s*인(?:용|승)?")


def parse_seat(name):
    m = _SEAT_COMBO_RE.search(name or "")
    if m:
        return f"{m.group(1)}인+{m.group(2)}인"
    seats = list(dict.fromkeys(_SEAT_RE.findall(name or "")))
    if not seats:
        return ""
    # 다중 인수 옵션 ("3인용 3.5인용") — 전부 보존
    return "/".join(f"{s}인용" for s in seats)


# 색상 판별 어휘 (긴 것 우선). 이 토큰을 포함해야 색상으로 인정.
_COLOR_VOCAB = sorted([
    "로즈골드", "샴페인골드", "멀티컬러", "아이보리", "라벤더", "버건디", "로즈",
    "화이트", "블랙", "그레이", "그레이지", "베이지", "브라운", "월넛", "오크",
    "네이비", "블루", "핑크", "옐로우", "그린", "레드", "크림", "카키",
    "올리브", "민트", "골드", "실버", "차콜", "와인", "퍼플", "오렌지",
    "내추럴", "우드", "체리", "메이플", "코랄", "머스타드", "테라코타",
    "먹색", "연청", "진청", "백색", "흑색", "회색",
], key=len, reverse=True)  # 주의: 바닥 "멀티"는 "멀티 인테리어" 오탐으로 제외
_COLOR_NOISE_RE = re.compile(
    r"language|한국어|english|구매\s*안\s*함|상세\s*페이지|선택|옵션|배송|스툴|쿠션|단품"
    r"|추가|증정|사은품|발받침|깔판"  # 추가구매 유도 옵션 제외 (2026-07-02)
    r"|\+\s*[\d,]+|\d+\s*[wW]\b|숏넥|롱넥|^\s*[A-Z0-9_-]{6,}\s*$", re.IGNORECASE)
_CCT_RE = re.compile(r"주광색|전구색|주백색|백색광|\d{4}[kK]\b")  # 색온도(광원색)


def norm_color(raw, is_lighting=False):
    parts = re.split(r"[|/,]", raw or "")
    out, seen = [], set()
    for v in parts:
        v = v.strip()
        v = re.sub(r"^(?:색상|컬러|color)\s*[:：]\s*", "", v, flags=re.IGNORECASE)
        v = re.sub(r"^\d{1,2}(?=[가-힣])", "", v)  # 옵션번호 프리픽스 "02아델그레이"
        if not v or _COLOR_NOISE_RE.search(v):
            continue
        if _CCT_RE.search(v):
            continue  # 색온도(주광색/전구색(오렌지빛) 등)는 광원색 — 전 카테고리에서 제외
        if not any(c in v for c in _COLOR_VOCAB):
            continue  # 색상 토큰 없는 옵션 문자열 제거
        v = lex.COLOR_KO.get(v.lower(), v)
        # 값 정리: 괄호류 제거·언더스코어→공백·NEW 접두 제거
        v = re.sub(r"^new[_\s]*", "", v, flags=re.IGNORECASE)
        v = re.sub(r"[()\[\]{}（）]", " ", v).replace("_", " ")
        v = re.sub(r"\s+", " ", v).strip(" -/,+")
        if not v:
            continue
        if v not in seen:
            seen.add(v)
            out.append(v)
    return " / ".join(out[:10])


# 소재 매칭: 긴 키 우선, 한 글자·모호 키("면"/"천"/"pu"/"pp")는 문맥 가드
_MAT_KEYS = sorted(lex.MATERIAL_MAP.items(), key=lambda kv: len(kv[0]), reverse=True)
_AMBIG_MAT = {"면", "천", "pu", "pp"}
_MYEON_OK_RE = re.compile(r"순면|면\s*100|면100|무명|광목|코튼|cotton", re.IGNORECASE)


def norm_material(raw):
    s = (raw or "").strip()
    if not s:
        return ""
    low = s.lower()
    for k, v in _MAT_KEYS:
        if k in _AMBIG_MAT:
            continue  # 모호 키는 아래 별도 처리
        if k in low:
            return v
    if _MYEON_OK_RE.search(low):
        return "면(코튼)"  # 표면/뒷면/단면의 '면' 오탐 방지
    return s


# ── 옵션 재수집 오버레이 (refetch_options.py 산출물 — 추출 CSV보다 우선) ──────
_OPT_CACHE = None


def load_options_overlay():
    global _OPT_CACHE
    if _OPT_CACHE is not None:
        return _OPT_CACHE
    import glob as _glob
    out = {}
    for fp in _glob.glob(os.path.join(OUT_DIR, "options_furniture_*.csv")):
        slug = os.path.basename(fp).replace("options_furniture_", "").replace(".csv", "")
        try:
            for r in csv.DictReader(open(fp, encoding="utf-8-sig")):
                if (r.get("options") or "").strip():
                    out[(slug, r["model_no"])] = r["options"]
        except Exception:
            continue
    _OPT_CACHE = out
    if out:
        print(f"[options] 재수집 옵션 오버레이 {len(out)}건 로드")
    return out


_GRP_CACHE = None


def load_option_groups_overlay():
    """옵션군 구조 오버레이(refetch_options --groups) — (slug, model_no) → JSON 문자열."""
    global _GRP_CACHE
    if _GRP_CACHE is not None:
        return _GRP_CACHE
    import glob as _glob
    out = {}
    for fp in _glob.glob(os.path.join(OUT_DIR, "options_groups_furniture_*.csv")):
        slug = os.path.basename(fp).replace("options_groups_furniture_", "").replace(".csv", "")
        try:
            for r in csv.DictReader(open(fp, encoding="utf-8-sig")):
                if (r.get("option_groups") or "").strip() not in ("", "[]"):
                    out[(slug, r["model_no"])] = r["option_groups"]
        except Exception:
            continue
    _GRP_CACHE = out
    if out:
        print(f"[options] 옵션군 오버레이 {len(out)}건 로드")
    return out


# ── OCR 고시 오버레이 (ocr_gosi_furniture.py 산출물) ─────────────────────────
_GOSI_CACHE = None


def load_gosi_overlay():
    """outputs/gosi_furniture_<slug>.csv → {(slug, model_no): {...}}"""
    global _GOSI_CACHE
    if _GOSI_CACHE is not None:
        return _GOSI_CACHE
    import glob as _glob
    out = {}
    for fp in _glob.glob(os.path.join(OUT_DIR, "gosi_furniture_*.csv")):
        slug = os.path.basename(fp).replace("gosi_furniture_", "").replace(".csv", "")
        try:
            for r in csv.DictReader(open(fp, encoding="utf-8-sig")):
                if any((r.get(k) or "").strip() for k in
                       ("material", "origin", "width_cm", "depth_cm", "height_cm", "safety_cert")):
                    out[(slug, r["model_no"])] = r
        except Exception:
            continue
    _GOSI_CACHE = out
    if out:
        print(f"[gosi] OCR 오버레이 {len(out)}건 로드")
    return out


def map_row(row):
    source = row.get("source", "")
    name = row.get("name", "")
    category = row.get("category", "")
    l1, l2 = classify(name, category, source)

    attrs = {
        # ── 메타 20 (근거 없으면 빈값 — CP-04) ──
        "product_type": l2 or "",
        "manufacturer": row.get("brand", ""),
        "brand_line": "",
        "model_no": row.get("model_no", ""),
        "color": norm_color(row.get("color", ""), is_lighting=(l1 == "조명")),
        # 침대사이즈 — 영문 코드 통일 (S/SS/D/Q/K, 2026-07-02 정책)
        "size": (row.get("bed_size", "")
                 if l1 in ("침실가구", "침구", "아동/주니어가구") else ""),
        "material": norm_material(row.get("material", "")),
        "width_cm": row.get("width_cm", ""),
        "depth_cm": row.get("depth_cm", ""),
        "height_cm": row.get("height_cm", ""),
        "net_weight_kg": "",
        "count_pcs": parse_count(name),
        "country_of_origin": row.get("origin", ""),
        "is_assembly_required": row.get("assembly", ""),
        "installation_service": row.get("installation_service", ""),
        "package_type": "",
        "safety_cert": row.get("safety_cert", ""),
        "warranty_period": "",
        "additive_free_claim": "",
        "key_features": "",
    }
    # ── 상품명 기반 보강 (스키마 추출소스 PRD_NM — 빈 값만 채움) ──
    if not attrs["color"]:
        attrs["color"] = name_color(name)
    if not attrs["material"]:
        attrs["material"] = name_material(name)
    if not (attrs["width_cm"] or attrs["depth_cm"] or attrs["height_cm"]):
        if l1 in ("거실가구", "침실가구", "주방가구", "수납가구",
                  "학생/사무용가구", "아동/주니어가구", "아웃도어가구"):
            w, d, h = name_three_dim(name)
            attrs["width_cm"], attrs["depth_cm"], attrs["height_cm"] = w, d, h
        elif l1 in ("침구", "카페트/러그", "커튼/블라인드", "홈패브릭/수예"):
            w, d = name_two_dim(name)
            attrs["width_cm"], attrs["depth_cm"] = w, d
            if w and d and not attrs["size"]:
                attrs["size"] = f"{w}x{d}"
        elif l1 == "인테리어소품":
            attrs["height_cm"] = name_height_cm(name)

    # ── 소재 축 정합성 가드 ──
    # 러그/매트: 라텍스·메모리폼은 뒷면 논슬립 백킹 — 상품명에 명시된 경우만 본체 소재로 인정
    if l1 in ("카페트/러그", "홈패브릭/수예") and attrs["material"] in ("라텍스", "메모리폼") \
            and attrs["material"] not in name:
        attrs["material"] = ""
    # 침구: 다운·메모리폼·라텍스는 충전재(filling) — material(겉감) 축에서 제거
    if l1 == "침구" and attrs["material"] in ("다운", "메모리폼", "라텍스"):
        attrs["material"] = ""

    # ── 대카 델타 파생 ──
    if l1 == "조명":
        attrs["lighting_type"] = first_match(name, LIGHTING_TYPE)
        attrs["light_source"] = first_match(name, LIGHT_SOURCE)
        attrs["size"] = ""  # 색온도(5700K 등) → 침대사이즈 오탐 방어
        # 스키마 중카 델타 (전구/조명 카탈로그 속성)
        attrs["light_color_temp"] = light_color_temp(name, row.get("color", ""))  # cat_41202
        pws = list(dict.fromkeys(_POWER_RE.findall(name)))
        attrs["power_w"] = "/".join(f"{p}W" for p in pws[:4])                     # cat_50108
        attrs["socket_count"] = _first(_SOCKET_RE, name, "구")                    # cat_49270
        fxs = list(dict.fromkeys(_FIXTURE_RE.findall(name)))
        attrs["fixture_count"] = "/".join(f"{f}등" for f in fxs[:4])              # cat_49264
        # 매입등·다운라이트·타공의 mm는 직경 — 길이 축과 분리
        if re.search(r"매입|다운라이트|타공|Ø", name):
            attrs["length_mm"] = ""
            attrs["ext_diameter_mm"] = _first(_LEN_MM_RE, name)
        else:
            attrs["length_mm"] = _first(_LEN_MM_RE, name)                         # cat_43293
            attrs["ext_diameter_mm"] = ""
        attrs["features"] = parse_features(name)                                  # cat_46459
        attrs["ext_size_inch"] = _first(_INCH_RE, name, "인치")
        ipm = _IP_RE.search(name)
        attrs["ext_ip_rating"] = f"IP{ipm.group(1)}" if ipm else ""
        attrs["ext_pyeong"] = _first(_PYEONG_RE, name, "평")
    elif l1 == "침구":
        bt = first_match(name, BEDDING_TYPE)
        attrs["bedding_type"] = bt
        # 커버·시트류는 충전재 없음 → filling 부여 금지
        attrs["filling_type"] = ("" if bt in _NO_FILLING or "커버" in name
                                 else first_match(f"{name} {row.get('material','')}", FILLING_TYPE))
        attrs["bedding_size"] = row.get("bed_size", "")
        # 스키마 subtype은 중카별 — 이불 하위유형은 l2=이불일 때만
        attrs["subtype"] = first_match(name, BEDDING_SUBTYPE) if l2 == "이불" else ""
        attrs["features"] = parse_features(name)                # cat_46459
        m = re.search(r"(\d{2,3})\s*수\b", name)
        attrs["ext_thread_count"] = f"{m.group(1)}수" if m else ""
    elif l1 == "거실가구":
        attrs["sofa_seat_count"] = parse_seat(name)             # cat_48263 사용인원
        finish, form, compo = sofa_attrs(name)
        attrs["finish_grade"] = finish                          # cat_48827 마감재
        attrs["sofa_form"] = form                               # cat_42446 형태
        attrs["composition"] = compo                            # cat_48438 구성
        attrs["features"] = parse_features(name)
    elif l1 == "인테리어소품":
        attrs["interior_item_type"] = first_match(name, INTERIOR_ITEM)
        if l2 == "크리스마스 트리":
            st = tree_subtype(name)
            # 트리 본체는 높이로 대형/미니 세분 (스키마 하위유형)
            h = attrs.get("height_cm")
            if st == "일반트리" and h and h.isdigit():
                if int(h) >= 210:
                    st = "대형트리"
                elif int(h) <= 90:
                    st = "미니트리"
            attrs["subtype"] = st
            m = re.search(r"장식\s*[:：]\s*(\d+)\s*개", name)
            attrs["ext_ornament_count"] = m.group(1) if m else ""
        elif l2 in ("디퓨저", "캔들"):
            attrs["scent"] = first_match(name, _SCENTS)         # cat_50873 향

    # OCR 고시 오버레이 — 빈 속성만 채움 (추출소스: PDP이미지OCR)
    g = load_gosi_overlay().get((source, row.get("model_no", "")))
    if g:
        for src_k, dst_k in (("material", "material"), ("origin", "country_of_origin"),
                             ("width_cm", "width_cm"), ("depth_cm", "depth_cm"),
                             ("height_cm", "height_cm"), ("safety_cert", "safety_cert")):
            v = (g.get(src_k) or "").strip()
            if v and not attrs.get(dst_k):
                attrs[dst_k] = norm_material(v) if dst_k == "material" else v

    # 참조 치수 가드 — "(추천트리:210cm)" 라벨 치수가 제품 높이로 오인 차단 (QA R5)
    # OCR 스펙 이미지가 추천 트리 크기를 보여주는 장식세트류에서 발생
    _ref = re.search(r"추천[^)]{0,10}[:：]\s*(\d{2,3})\s*cm", name)
    if _ref and attrs.get("height_cm") == _ref.group(1):
        attrs["height_cm"] = ""

    # 옵션 재수집 오버레이 — 추출 시 놓친 드롭다운 우선 사용
    _ropt = load_options_overlay().get((source, row.get("model_no", "")))
    raw_opts = _ropt if _ropt else row.get("color", "")
    if _ropt and not attrs["color"]:
        attrs["color"] = norm_color(_ropt.replace("|", " / "), is_lighting=(l1 == "조명"))

    return {
        "prd_id": f"{source}_{row.get('model_no','')}",
        "raw_options": raw_opts,
        "raw_option_groups": load_option_groups_overlay().get((source, row.get("model_no", "")), ""),
        "meta_category": "가구/인테리어",
        "l1_category": l1,
        "l2_category": l2,
        "name": name,
        "price": row.get("price", ""),
        "attributes": attrs,
        "source": {"mall": source, "url": row.get("url", ""),
                   "extract_source": "공식몰PDP"},
    }


VARIANTS_OUT = os.path.join(OUT_DIR, "furniture_geo_variants.jsonl")


# 침구 사이즈 옵션: 상품명 "[S/SS,Q/K]" — 콤마가 옵션 구분, "S/SS"는 겸용 1옵션
_BED_SIZE_OPT_RE = re.compile(r"\[([SQKD/,\s]+)\]")
_SIZE_TOKEN_RE = re.compile(r"^(?:S|SS|D|Q|K|KK)(?:/(?:S|SS|D|Q|K|KK))*$")


def bedding_size_options(name, fallback=""):
    """이불 상품명에서 사이즈 옵션 목록. 없으면 fallback 단일 or 빈 1개."""
    m = _BED_SIZE_OPT_RE.search(name or "")
    if m:
        opts = [t.strip() for t in m.group(1).split(",") if t.strip()]
        opts = [t for t in opts if _SIZE_TOKEN_RE.match(t)]
        if opts:
            return opts
    return [fallback] if fallback else [""]


# ── 옵션 드롭다운 → 사이즈 보조 전개 (상품명에 사이즈 없을 때) ────────────────
# 최장일치 우선 — '슈퍼킹'이 '킹'으로, '멀티싱글'이 '싱글'로 오치환되지 않게
_SIZE_KO2CODE = [("킹오브킹", "KOK"), ("엑스라지킹", "XLK"), ("패밀리킹", "FK"),
                 ("슈퍼싱글", "SS"), ("슈퍼킹", "SK"), ("멀티싱글", "MS"),
                 ("라지킹", "LK"), ("이스턴킹", "EK"),
                 ("싱글", "S"), ("더블", "D"), ("퀸", "Q"), ("킹", "K")]
# 사이즈 단어를 부분문자열로 포함하는 색상어 — 치환 전 마스킹 ('파우더블루'→'더블')
_SIZE_COLOR_MASK = re.compile(r"파우더블루")
# 한글/영문/숫자 인접 배제("D형"·"K5"·"S자") + '-숫자'("D-1타입")·'.'("D-01.") 열거자 배제
_OPT_SIZE_TOKEN = re.compile(
    r"(?<![A-Za-z0-9가-힣])(KOK|XLK|SS|SK|MS|EK|LK|KK|FK|[SQKD])"
    r"(?!(?:[A-Za-z0-9가-힣]|\s*-\s*\d|\.))")
_OPT_SIZE_SKIP = re.compile(r"추가|사은품|증정|선택\s*안\s*함|구매\s*안\s*함|미포함|제외")
# 추가구매/사은품 섹션 헤더 — 이후 라벨 전체 컷 (본 옵션은 항상 헤더보다 앞)
_OPT_SECTION_END = re.compile(r"추가\s*상품|추가\s*구성품?|사은품")


def _size_norm(label):
    t = _SIZE_COLOR_MASK.sub(" ", label)
    for ko, cd in _SIZE_KO2CODE:
        t = t.replace(ko, f" {cd} ")
    return t


def _size_extract(label, out):
    """라벨 1개에서 사이즈 값 추출해 out에 추가.
    '+'/'/'로 이어진 복수 토큰은 조합 1값("SS+Q" 패밀리, "S/SS" 겸용).
    인접 중복 토큰은 병기('슈퍼싱글(SS)')/직결합('퀸Q')이므로 1개로 접기.
    떨어진 이종 토큰은 사이즈 축이 아닌 서술 — 라벨 스킵."""
    t = _size_norm(label)
    ms = list(_OPT_SIZE_TOKEN.finditer(t))
    if not ms:
        return
    val, prev = ms[0].group(1), ms[0]
    for b in ms[1:]:
        sep = t[prev.end():b.start()].strip()
        if b.group(1) == prev.group(1) and sep not in ("+", "/"):
            prev = b
            continue
        if sep in ("+", "/"):
            val += sep + b.group(1)
            prev = b
        else:
            return
    if val not in out:
        out.append(val)


def bed_sizes_from_options(raw_options):
    """드롭다운 옵션에서 사이즈 축 추출 (상품명에 사이즈 없을 때 보조 전개).
    vittz류 flatten 옵션('|' 없이 '/' 나열, 색상 조각 혼재)은 조각별 추출."""
    out = []
    raw = raw_options or ""
    single = "|" not in raw
    for label in raw.split("|"):
        label = label.strip()
        if label and _OPT_SECTION_END.fullmatch(label):
            break
        if not label or _OPT_SIZE_SKIP.search(label):
            continue
        pieces = [p.strip() for p in label.split("/") if p.strip()]
        if single and len(pieces) >= 4 and \
                any(not _OPT_SIZE_TOKEN.search(_size_norm(p)) for p in pieces):
            for p in pieces:      # flatten 나열 — '+' 조합은 조각 내 유지
                _size_extract(p, out)
        else:
            _size_extract(label, out)
    return out if 1 <= len(out) <= 8 else []


def expand_variants(records):
    """변형 전개 — 부모(PDP) 1건 → 변형 N건.
    변형축: 색상 (전 카테고리) × 사이즈 (침구만 — 상품명 [S/SS,Q/K] 옵션).
    스포츠 catalog_decomposed(사이즈 전개)와 대칭."""
    out = []
    for r in records:
        colors = [c.strip() for c in (r["attributes"].get("color") or "").split(" / ") if c.strip()]
        if not colors:
            colors = [""]
        if r["l1_category"] == "침구":
            sizes = bedding_size_options(r["name"], r["attributes"].get("bedding_size", ""))
            if sizes == [""]:  # 이름·속성에 없으면 옵션 드롭다운에서 보조 전개
                sizes = bed_sizes_from_options(r.get("raw_options", "")) or [""]
        elif r["l1_category"] == "침실가구":
            # 침대/매트리스: 상품명 나열 "S/SS/Q/K" + 한글 표기 → 영문 코드로 전개
            codes = re.findall(r"(?<![A-Za-z])(SS|EK|LK|KK|[SQKD])(?![A-Za-z])", r["name"])
            t = _SIZE_COLOR_MASK.sub(" ", r["name"])  # '파우더블루'→'더블' 오탐 방지
            for ko, cd in _SIZE_KO2CODE:  # 최장일치 우선 공유 어휘 ('슈퍼킹'≠'킹')
                if ko in t:
                    codes.append(cd)
                    t = t.replace(ko, " ")
            codes = list(dict.fromkeys(codes))
            if not codes:  # 이름에 없으면 옵션 드롭다운에서 보조 전개
                codes = bed_sizes_from_options(r.get("raw_options", ""))
            sizes = codes if codes else ([r["attributes"].get("size", "")] if r["attributes"].get("size") else [None])
        else:
            sizes = [None]  # 사이즈축 미적용
        for c in colors:
            for sz in sizes:
                v_attrs = dict(r["attributes"])
                v_attrs["color"] = c
                if sz is not None:
                    v_attrs["size"] = sz
                cslug = re.sub(r"[^A-Za-z0-9가-힣]", "", c)[:24] or "nocolor"
                sslug = ("_" + re.sub(r"[^A-Za-z0-9]", "", sz)) if sz else ""
                axes = c + (f" / {sz}" if sz else "")
                out.append({
                    "variant_id": f"{r['prd_id']}_{cslug}{sslug}",
                    "parent_uid": r["prd_id"],
                    "meta_category": r["meta_category"],
                    "l1_category": r["l1_category"],
                    "l2_category": r["l2_category"],
                    "name": r["name"],
                    "variant_value": axes,        # 색상 [/ 사이즈]
                    "variant_color": c,
                    "variant_size": sz or "",
                    "price": r["price"],
                    "attributes": v_attrs,
                    "source": r["source"],
                })
    return out


def build_report(records):
    """대카별 속성 채움률 md 리포트."""
    from collections import defaultdict
    by_l1 = defaultdict(list)
    for r in records:
        by_l1[r["l1_category"] or "(미분류)"].append(r)

    lines = ["# 가구/인테리어 GEO 매핑 리포트", "",
             f"총 {len(records)}개 상품 · 소스: furniture_all_brands.csv", "",
             "## 대카 분포", "", "| 대카 | 상품수 | 중카 분류율 |", "|---|---:|---:|"]
    for l1, rs in sorted(by_l1.items(), key=lambda x: -len(x[1])):
        l2_rate = sum(1 for r in rs if r["l2_category"]) / len(rs) * 100
        lines.append(f"| {l1} | {len(rs)} | {l2_rate:.0f}% |")

    lines += ["", "## 대카별 속성 채움률 (메타 주요 + 델타)", ""]
    KEY_ATTRS = ["product_type", "manufacturer", "model_no", "color", "size",
                 "material", "width_cm", "country_of_origin", "safety_cert",
                 "count_pcs", "lighting_type", "light_source", "bedding_type",
                 "filling_type", "sofa_seat_count", "interior_item_type"]
    for l1, rs in sorted(by_l1.items(), key=lambda x: -len(x[1])):
        lines.append(f"### {l1} ({len(rs)}개)")
        lines.append("")
        lines.append("| 속성 | 채움 | 비율 |")
        lines.append("|---|---:|---:|")
        for a in KEY_ATTRS:
            vals = [r["attributes"].get(a, "") for r in rs]
            if all(v == "" for v in vals) and a not in ("product_type", "manufacturer"):
                applicable = any(a in r["attributes"] for r in rs)
                if not applicable:
                    continue
            n = sum(1 for v in vals if v)
            if n == 0 and a in ("lighting_type", "light_source", "bedding_type",
                                "filling_type", "sofa_seat_count", "interior_item_type"):
                continue
            lines.append(f"| `{a}` | {n}/{len(rs)} | {n/len(rs)*100:.0f}% |")
        lines.append("")
    return "\n".join(lines)


def write_xlsx(records, variants=None):
    """엑셀 출력 — 시트: GEO매핑(플랫) / 색상변형 / 채움률 / 대카분포."""
    from collections import defaultdict
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter

    META_COLS = ["product_type", "manufacturer", "brand_line", "model_no", "color",
                 "size", "material", "width_cm", "depth_cm", "height_cm",
                 "net_weight_kg", "count_pcs", "country_of_origin",
                 "is_assembly_required", "installation_service", "package_type",
                 "safety_cert", "warranty_period", "key_features"]
    DELTA_COLS = ["lighting_type", "light_source", "light_color_temp", "power_w",
                  "socket_count", "fixture_count", "length_mm",
                  "bedding_type", "filling_type", "bedding_size", "subtype",
                  "sofa_seat_count", "finish_grade", "sofa_form", "composition",
                  "interior_item_type", "scent", "features",
                  "ext_size_inch", "ext_ip_rating", "ext_pyeong",
                  "ext_thread_count", "ext_ornament_count", "ext_diameter_mm"]

    wb = Workbook()
    head_font = Font(bold=True, color="FFFFFF")
    head_fill = PatternFill("solid", fgColor="4472C4")

    # ── 시트 1: GEO매핑 ──
    ws = wb.active
    ws.title = "GEO매핑"
    headers = (["prd_id", "쇼핑몰", "대카", "중카", "상품명", "가격"]
               + META_COLS + DELTA_COLS + ["url"])
    ws.append(headers)
    for c in ws[1]:
        c.font, c.fill = head_font, head_fill
    for r in records:
        a = r["attributes"]
        ws.append([r["prd_id"], r["source"]["mall"], r["l1_category"],
                   r["l2_category"], r["name"], int(r["price"]) if str(r["price"]).isdigit() else r["price"]]
                  + [a.get(k, "") for k in META_COLS]
                  + [a.get(k, "") for k in DELTA_COLS]
                  + [r["source"]["url"]])
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for i, w in enumerate([18, 10, 12, 12, 50, 11], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # ── 시트 1.5: 색상변형 (identity 단위) ──
    if variants:
        wsv = wb.create_sheet("색상변형")
        wsv.append(["variant_id", "parent_uid", "쇼핑몰", "대카", "중카", "상품명",
                    "색상", "사이즈", "가격", "material", "url"])
        for c in wsv[1]:
            c.font, c.fill = head_font, head_fill
        for v in variants:
            a = v["attributes"]
            wsv.append([v["variant_id"], v["parent_uid"], v["source"]["mall"],
                        v["l1_category"], v["l2_category"], v["name"],
                        v.get("variant_color", ""), v.get("variant_size", ""),
                        int(v["price"]) if str(v["price"]).isdigit() else v["price"],
                        a.get("material", ""), v["source"]["url"]])
        wsv.freeze_panes = "A2"
        wsv.auto_filter.ref = wsv.dimensions
        for i, w in enumerate([26, 18, 10, 12, 12, 46, 14, 11], 1):
            wsv.column_dimensions[get_column_letter(i)].width = w

    # ── 시트 2: 채움률 ──
    ws2 = wb.create_sheet("채움률")
    by_l1 = defaultdict(list)
    for r in records:
        by_l1[r["l1_category"] or "(미분류)"].append(r)
    ws2.append(["대카", "상품수"] + META_COLS + DELTA_COLS)
    for c in ws2[1]:
        c.font, c.fill = head_font, head_fill
    for l1, rs in sorted(by_l1.items(), key=lambda x: -len(x[1])):
        row = [l1, len(rs)]
        for k in META_COLS + DELTA_COLS:
            n = sum(1 for r in rs if r["attributes"].get(k))
            row.append(f"{n/len(rs)*100:.0f}%" if n else "")
        ws2.append(row)
    ws2.freeze_panes = "C2"
    ws2.column_dimensions["A"].width = 14

    # ── 시트 3: 대카분포 ──
    ws3 = wb.create_sheet("대카분포")
    ws3.append(["대카", "상품수", "중카 분류율", "쇼핑몰"])
    for c in ws3[1]:
        c.font, c.fill = head_font, head_fill
    for l1, rs in sorted(by_l1.items(), key=lambda x: -len(x[1])):
        l2r = sum(1 for r in rs if r["l2_category"]) / len(rs) * 100
        malls = sorted({r["source"]["mall"] for r in rs})
        ws3.append([l1, len(rs), f"{l2r:.0f}%", ", ".join(malls)])
    ws3.column_dimensions["A"].width = 14
    ws3.column_dimensions["D"].width = 40

    wb.save(XLSX_OUT)
    print(f"[xlsx] {len(records)}행 → {XLSX_OUT}")


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--in", dest="in_path", default=IN_DEFAULT)
    ap.add_argument("--report-only", action="store_true")
    args = ap.parse_args()

    if args.report_only and os.path.exists(JSONL_OUT):
        records = [json.loads(l) for l in open(JSONL_OUT, encoding="utf-8")]
    else:
        records = []
        with open(args.in_path, encoding="utf-8-sig", newline="") as f:
            for row in csv.DictReader(f):
                if row.get("name"):
                    records.append(map_row(row))
        with open(JSONL_OUT, "w", encoding="utf-8") as f:
            for r in records:
                f.write(json.dumps(r, ensure_ascii=False) + "\n")
        print(f"[map] {len(records)}건 → {JSONL_OUT}")

    variants = expand_variants(records)
    with open(VARIANTS_OUT, "w", encoding="utf-8") as f:
        for v in variants:
            f.write(json.dumps(v, ensure_ascii=False) + "\n")
    print(f"[variants] 색상 전개 {len(records)} → {len(variants)}건 → {VARIANTS_OUT}")

    report = build_report(records)
    with open(REPORT_OUT, "w", encoding="utf-8") as f:
        f.write(report + "\n")
    print(f"[report] → {REPORT_OUT}")

    write_xlsx(records, variants)


if __name__ == "__main__":
    main()
